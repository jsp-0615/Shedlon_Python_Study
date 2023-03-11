"""
@author:jsp
@file:python_excel_demo_02.py
@time:2023/3/11 19:13
@desc:
"""
import os
from openpyxl import load_workbook
file_path=os.path.join("J:\\","tempDocument","颜赵刚_发票整理.xlsx")
# print(file_path)
if os.path.exists(file_path):
    # [<Worksheet "发票">, <Worksheet "机械作业">] 每个worksheet相当于一个个对象，里面有很多单元格
    invoice=load_workbook(file_path)
    # print(invoice.worksheets)
    #sheet1=invoice.wokrsheets[0]
else:
    print("文件不存在")
# print(invoice.worksheets[0])
#拿到表格中的总价的所有内容
sheet1 = invoice.worksheets[0]
# print(sheet1)
for row in sheet1.iter_rows(min_row=2):# row是那一行的数据（cell对象，cell对象，cell对象，cell对象，cell对象.....)
    cell = row[1]
    print(cell.value)
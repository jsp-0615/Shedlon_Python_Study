"""
@author:jsp
@file:excel_to_database.py
@time:2023/3/8 16:23
@desc:
"""
import pyodbc
from openpyxl import load_workbook


class Emp(object):
    id = ""
    name = ""
    age = ""
    deptId = ""
    empno = ""
    hire_date = ""
    salary = ""
    job_description = ""
    comm = ""

    def __init__(self, data):
        self.id = data.get('id')
        self.name = data.get('name')
        self.age = data.get('age')
        self.deptId = data.get('deptId')
        self.empno = data.get('empno')
        self.hire_date = data.get('hire_date')
        self.salary = data.get('salary')
        self.job_description = data.get('job_description')
        self.comm = data.get('comm')


server = 'DESKTOP-DH171PS\SQLSERVERSHELDON'  # SQL Server服务器名称
database = 'DB_ORAPS'  # 数据库名称
username = 'sa'  # 数据库用户名
password = '0615..jsp.'  # 数据库密码
cnxn = pyodbc.connect(
    'DRIVER={SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
cursor = cnxn.cursor()
excel = load_workbook("J:\\tempDocument\\员工表格.xlsx")
activeSheet = excel["帮派"]
# 先获取第一行我设置的列名，同时这里应该和数据库的列名一致，方便等会做一致性检测
# print(activeSheet.max_column)
# print(activeSheet.max_row)
varialeName = []
for row in activeSheet.iter_rows(min_row=1, max_row=1):
    for cell in row:
        varialeName.append(cell.value)
# print(varialeName)
dict = {}
emps = []
# 自己去给参数应该从第几行读到第几行，但对于列可以直接通过参数赋值
for row in activeSheet.iter_rows(min_row=2, max_row=6, max_col=activeSheet.max_column):
    i = 0
    for cell in row:
        dict[varialeName[i]] = cell.value
        i += 1
    emp = Emp(dict)

    emps.append(emp)  # 为了避免反复落一次表连一次把它放到一个数组中，一次性的插入表中
#     cursor.execute('insert into user (id, name) values (%s, %s)', ['1', 'Michael'])
i=0
for emp in emps:
    sql_insert = "insert into t_emp(id,name,age,deptId,empno,hire_date,salary,job_description,comm) values (?,?,?,?,?,?,?,?,?)"
    # print(sql_insert)
    cursor.execute(sql_insert,
                       emp.id, emp.name, emp.age, emp.deptId, emp.empno, emp.hire_date, emp.salary, emp.job_description,
                       emp.comm)
    print(emp.name)
cnxn.commit()
cursor.close()
cnxn.close

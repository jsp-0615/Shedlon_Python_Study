from openpyxl import load_workbook
#实例化
#wb=Workbook()#create an excel file in RAM
#获取当前active的sheet
#as=wb.active
#打印sheet表名
#print(ws.title)
#改sheet名
#ws.title="通过python改名"
#保存
#wb.save("excel_python.xlsx")

# 写数据
# 数据一 数据可以直接分配到单元格中（可以输入公式
#as["B6"]="蒋天生"
# 方式二：可以附加行，从第一列开始附加（从最下方空白处，最左开始）（可以输入多行)
#as.append([1,2,3])
#这样子写的话 就会从有数据的最后一行的下一行开始从A列开始写1，B列写2，C列写3
# 方式三： python类型会被自动转换
# excel=load_workbook("J:\\tempDocument\\员工表格.xlsx")
# sheetName="员工表"
# activcSheet=excel[sheetName]
# print(activcSheet["B3:B5"])
# for cell in activcSheet["B1:B5"]:
#     print(cell[0].value)
# 获取表内的所有数据
#
# for row in activcSheet:
#     for cell in row:
#         print(cell.value,end=",")
#     print()
# 按列遍历
# 像A1，A2，A3这样的顺序
# for column in activcSheet.columns:
#     for cell in column:
#         print(cell.value,end=",")
#     print()

# 遍历指定行&列
#从第二行开始到第五行，每行打印第二列开始，打印指定列
# for row in activcSheet.iter_rows(min_row=1,max_row=6,min_col=2):
#     for cell in row:
#         print(cell.value,end=",")
#     print()

#按指定列
# for cols in activcSheet.iter_cols(min_col=2,max_col=3):
#     for cell in cols:
#         print(cell.value, end=",")
#     print()

#删除工作表
#方法一：
# excel.remove(activcSheet)
# #方法二
# del excel(activcSheet)

# empInfo=["蒋天生",45,10,"100015","1999/06/01",199999,"大佬"]
# i=0
# for col in ["B","C","D","E","F","G","H"]:
#     activcSheet[col+"6"]=empInfo[i]
#     i=i+1
# excel.save("J:\\tempDocument\\员工表格.xlsx")